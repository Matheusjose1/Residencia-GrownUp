import os
import openpyxl
from openpyxl.styles import Font, Alignment

from app.core.config import XLSX_RESULTS_DIR  # CSV_RESULTS_DIR removido


def extract_id_from_filename(filename: str) -> str | None:
    """
    Extrai um ID de 4 dígitos do nome do arquivo.
    """
    name_without_ext = os.path.splitext(os.path.basename(filename))[0]
    numeros = ''.join(filter(str.isdigit, name_without_ext))
    return numeros[-4:] if len(numeros) >= 4 else None


def save_detection_to_xlsx(detection_data: list[dict], filename: str) -> str:
    """
    Salva os dados de detecção de objetos em um arquivo XLSX com os campos:
    ID (do nome da imagem), Tipo (classe detectada), Acurácia (confiança).

    detection_data = [
        {"image_id": "0001", "class": "domiciliar", "confidence": 0.95},
        {"image_id": "0001", "class": "volumoso", "confidence": 0.88},
        {"image_id": "0002", "class": "poda", "confidence": 0.70},
        ...
    ]
    """
    output_path = XLSX_RESULTS_DIR / filename

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Detecções YOLO"

    # Definir cabeçalhos conforme solicitado
    headers = ["ID", "Tipo", "Acurácia"]
    ws.append(headers)

    # Estilizar cabeçalhos
    for col_idx, header_text in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header_text)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 20

    for det_item in detection_data:
        image_id = det_item.get("image_id", "N/A")
        class_name = det_item.get("class", "N/A")
        confidence = det_item.get("confidence", 0.0)

        ws.append([image_id, class_name, f"{confidence * 100:.2f}%"])  # Acurácia em porcentagem

    # Ajustar largura das colunas automaticamente
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 2

    wb.save(output_path)
    return str(output_path)